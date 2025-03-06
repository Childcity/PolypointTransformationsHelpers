from RBF_exosceleton.obj_io import *
import numpy as np
import math
import time
import collections
from scipy.optimize import minimize
from enum import Enum
from datetime import datetime
from openpyxl import Workbook, load_workbook

reg_term = 1e-16


class Plane:
	_nan_p = np.array([np.nan, np.nan, np.nan])

	id = -1

	A, B, C, D = np.nan, np.nan, np.nan, np.nan
	P1, P2, P3 = _nan_p, _nan_p, _nan_p

	@staticmethod
	def from_abcd(id, a, b, c, d):
		p = Plane()
		p.id = id
		p.A, p.B, p.C, p.D = a, b, c, d
		return p

	def __init__(self, id = -1, p1 = _nan_p, p2 = _nan_p, p3 = _nan_p):
		if np.any(np.isnan(p1)) or np.any(np.isnan(p2)) or np.any(np.isnan(p3)):
			return

		self.id = id
		self.P1, self.P2, self.P3 = p1, p2, p3

		# Calculate two vectors in the plane
		p2p1 = p2 - p1
		p3p1 = p3 - p1

		# Calculate the normal vector to the plane using the cross product
		normal_vector = np.cross(p2p1, p3p1)

		# Get plane coefficients
		self.A, self.B, self.C = normal_vector

		# Calculate the value of D for the plane equation ax + by + cz + d = 0
		self.D = -(self.A * p1[0] + self.B * p1[1] + self.C * p1[2])

	def orthogonal_plane(self, point, other_vector, id = None):
		this_plane_normal_vector = np.array([self.A, self.B, self.C])
		other_plane_normal_vector = np.cross(this_plane_normal_vector, other_vector)
		a, b, c = other_plane_normal_vector
		d = -(a * point[0] + b * point[1] + c * point[2])
		return Plane.from_abcd(id, a, b, c, d)

	def __repr__(self):
		return f"Plane(id={self.id}; A={self.A}; B={self.B}; C={self.C}; D={self.D})"

	def __str__(self):
		return f"Plane(id={self.id}; A={self.A}; B={self.B}; C={self.C}; D={self.D})"

	def get_surface_for_z(self, X, Y): # X, Y: 2D arrays (2D grid)
		return (-self.A * X - self.B * Y - self.D) / self.C
 
	def length(self):
		# magnitude of the normal vector
		return math.sqrt(self.A ** 2 + self.B ** 2 + self.C ** 2)

	def normalized(self):
		magnitude = self.length()
		if magnitude == 0:
			return self
		else:
			return Plane.from_abcd(self.id, self.A / magnitude, self.B / magnitude, self.C / magnitude, self.D / magnitude)

	def sign_distance(self, point):
		A, B, C, D = self.A, self.B, self.C, self.D
		x, y, z = point
		return (A * x + B * y + C * z + D) / self.length()

	def distance(self, point):
		A, B, C, D = self.A, self.B, self.C, self.D
		x, y, z = point
		return (A * x + B * y + C * z + D) ** 2 / self.length()


class ElapsedTime:
	def __init__(self):
		self.start_time = time.time_ns()

	def elapsed(self, round_to=1):
		elapsed_time = (time.time_ns() - self.start_time) / 1e6
		return round(elapsed_time, round_to)


Vertex = collections.namedtuple('Vertex', ['x', 'y', 'z'])
PlaneSet = set[Plane]
PlanesForVertex = dict[Vertex, PlaneSet]
TriPlane = tuple[Plane, Plane, Plane]
TriPlaneSet = set[TriPlane]
TriPlaneForVertex = dict[Vertex, TriPlane]
TriPlanesForVertex = dict[Vertex, TriPlaneSet]
Topology = Enum(value='Topology', names=('Intersect Sidor Orthogonal'))

exec_stat: dict[str, int] = {}

def build_planes_intersect_topology(vertexes: list, triangles: list) -> tuple[list[Plane], PlanesForVertex]:
	planes: list[Plane] = []
	planes_for_vertex_dict: PlanesForVertex = {
		Vertex(*v_arr): set() for v_arr in vertexes # Initialize with empty sets, to preserve order of 'vertexes'
	}

	# NOTE: planes_for_vertex_dict stores each vertex as list of connected planes 

	def append_vertex_as_planes(point: list, plane: Plane):
		vertex_key = Vertex(*point)
		assert vertex_key in planes_for_vertex_dict
		planes_for_vertex_dict[vertex_key].add(plane)

	# Generate planes for each triangle
	for id, tri  in enumerate(triangles):
		p1 = np.array(triangle_point(vertexes, tri[0]))
		p2 = np.array(triangle_point(vertexes, tri[1]))
		p3 = np.array(triangle_point(vertexes, tri[2]))
		plane = Plane(id, p1, p2, p3)
		planes.append(plane)
		append_vertex_as_planes(p1, plane)
		append_vertex_as_planes(p2, plane)
		append_vertex_as_planes(p3, plane)
  
	return planes, planes_for_vertex_dict


def build_planes_sidor_topology(vertexes: list, triangles: list) -> tuple[list[Plane], TriPlanesForVertex]:
	planes: list[Plane] = []
	planes_for_vertex_dict: TriPlanesForVertex = {
		Vertex(*v_arr): set() for v_arr in vertexes # Initialize with empty sets, to preserve order of 'vertexes'
	}

	# NOTE: planes_for_vertex_dict stores each vertex as list of connected planes 

	def append_vertex_as_tri_plane(point: list, tri_plane: TriPlane):
		vertex_key = Vertex(*point)
		assert vertex_key in planes_for_vertex_dict
		planes_for_vertex_dict[vertex_key].add(tri_plane)

	# Generate planes for each triangle
	plane_id = -1
	for tri in triangles:
		p1 = np.array(triangle_point(vertexes, tri[0]))
		p2 = np.array(triangle_point(vertexes, tri[1]))
		p3 = np.array(triangle_point(vertexes, tri[2]))

		triangle_plane = Plane(plane_id + 1, p1, p2, p3).normalized()
		p1p2_plane = triangle_plane.orthogonal_plane(p1, p2 - p1, plane_id + 2).normalized()
		p2p3_plane = triangle_plane.orthogonal_plane(p2, p3 - p2, plane_id + 3).normalized()
		p3p1_plane = triangle_plane.orthogonal_plane(p3, p1 - p3, plane_id + 4).normalized()

		planes.append(triangle_plane)	# inserted at plane_id + 1
		planes.append(p1p2_plane)		# inserted at plane_id + 2
		planes.append(p2p3_plane)		# inserted at plane_id + 3
		planes.append(p3p1_plane)		# inserted at plane_id + 4

		# In Sidor diser 1 vertex represented as 3 planes (TriPlane)
		append_vertex_as_tri_plane(p1, (triangle_plane, p1p2_plane, p3p1_plane))
		append_vertex_as_tri_plane(p2, (triangle_plane, p1p2_plane, p2p3_plane))
		append_vertex_as_tri_plane(p3, (triangle_plane, p2p3_plane, p3p1_plane))

		plane_id += 4

	assert plane_id == len(planes) - 1
	return planes, planes_for_vertex_dict


def build_planes_orthogonal_topology(vertexes: list, triangles: list) -> tuple[list[Plane], TriPlaneForVertex]:
	planes: list[Plane] = []
	planes_for_vertex_dict: TriPlaneForVertex = {
		Vertex(*v_arr): TriPlane() for v_arr in vertexes # Initialize with empty TriPlane, to preserve order of 'vertexes'
	}

	# NOTE: planes_for_vertex_dict stores each vertex as TriPlane (3 orthogonal planes)

	def append_vertex_as_tri_plane(point: list, tri_plane: TriPlane):
		vertex_key = Vertex(*point)
		assert vertex_key in planes_for_vertex_dict
		planes_for_vertex_dict[vertex_key] = tri_plane
		assert len(planes_for_vertex_dict[vertex_key]) == 3

	# Generate planes for each triangle
	plane_id = -1
	for tri in triangles:
		p1 = np.array(triangle_point(vertexes, tri[0]))
		p2 = np.array(triangle_point(vertexes, tri[1]))
		p3 = np.array(triangle_point(vertexes, tri[2]))

		triangle_plane = Plane(plane_id + 1, p1, p2, p3).normalized()
		p1p2_plane = triangle_plane.orthogonal_plane(p1, p2 - p1, plane_id + 2).normalized()
		p2p3_plane = triangle_plane.orthogonal_plane(p2, p3 - p2, plane_id + 3).normalized()
		p3p1_plane = triangle_plane.orthogonal_plane(p3, p1 - p3, plane_id + 4).normalized()

		planes.append(triangle_plane)	# inserted at plane_id + 1
		planes.append(p1p2_plane)		# inserted at plane_id + 2
		planes.append(p2p3_plane)		# inserted at plane_id + 3
		planes.append(p3p1_plane)		# inserted at plane_id + 4

		if planes_for_vertex_dict[Vertex(*p1)] == TriPlane():
			append_vertex_as_tri_plane(p1, (triangle_plane, p1p2_plane, p3p1_plane))

		if planes_for_vertex_dict[Vertex(*p2)] == TriPlane():
			append_vertex_as_tri_plane(p2, (triangle_plane, p1p2_plane, p2p3_plane))

		if planes_for_vertex_dict[Vertex(*p3)] == TriPlane():
			append_vertex_as_tri_plane(p3, (triangle_plane, p2p3_plane, p3p1_plane))

		plane_id += 4

	assert plane_id == len(planes) - 1
	return planes, planes_for_vertex_dict


def build_planes(vertexes: list, triangles: list, topology: Topology) -> tuple[list[Plane], PlanesForVertex] | tuple[list[Plane], TriPlanesForVertex]:
	match topology:
		case Topology.Intersect:
			return build_planes_intersect_topology(vertexes, triangles)
		case Topology.Sidor:
			return build_planes_sidor_topology(vertexes, triangles)
		case Topology.Orthogonal:
			return build_planes_orthogonal_topology(vertexes, triangles)
	assert False


def get_polypoint_plane(plane: Plane, orig_basises: list, res_basises: list):
	a1, b1, c1, d1, r1 = 0, 0, 0, 0, 0
	a2, b2, c2, d2, r2 = 0, 0, 0, 0, 0
	a3, b3, c3, d3, r3 = 0, 0, 0, 0, 0
	a4, b4, c4, d4, r4 = 0, 0, 0, 0, 0

	for orig_basis_p, res_basis_p in zip(orig_basises, res_basises):
		γ = plane.sign_distance(orig_basis_p)
		# print("orig_basis_p", orig_basis_p, "\t γ", γ)

		x = res_basis_p[0]
		y = res_basis_p[1]
		z = res_basis_p[2]
		# h = 1
		γSquered = γ ** 2

		a1 += x * x / γSquered
		b1 += x * y / γSquered
		c1 += x * z / γSquered
		d1 += x 	/ γSquered		# * h

		# a2 += y * x / γSquered	# a2 == b1
		b2 += y * y	/ γSquered
		c2 += y * z	/ γSquered
		d2 += y 	/ γSquered		# * h

		#a3 += z * x * γSquered		# a3 == c1
		#b3 += z * y * γSquered		# b3 == c2
		c3 += z * z	/ γSquered
		d3 += z 	/ γSquered		# * h

		#a4 += x / γSquered		# a4 == d1
		#b4 += y / γSquered		# b4 == d2
		#c4 += z / γSquered		# c4 == d3
		d4 += 1 / γSquered 		# * h * h
  
		r1 += x / γ
		r2 += y / γ
		r3 += z / γ
		r4 += 1 / γ

	# Add a small regularization term to the diagonal elements of A
	A = np.array([
		[a1 + reg_term, b1, c1, d1],
		[b1, b2 + reg_term, c2, d2],
		[c1, c2, c3 + reg_term, d3],
		[d1, d2, d3, d4 + reg_term]])

	B = np.array([r1, r2, r3, r4])
	X = np.linalg.solve(A, B)
	return Plane.from_abcd(plane.id, X[0], X[1], X[2], X[3])


def get_polypoint_planes_list(in_planes: list[Plane], orig_basises: list, res_basises: list) -> list[Plane]:
	elapsed_time = ElapsedTime()
	result_planes: list = list()

	for plane in in_planes:
		tr_plane = get_polypoint_plane(plane, orig_basises, res_basises) # .normalized()
		assert tr_plane.id != -1
		result_planes.append(tr_plane)

	exec_stat["get_polypoint_planes_list"] = elapsed_time.elapsed()
	return result_planes


def closest_point_to_planes_min(planes: PlaneSet) -> np.array:
	def objective_function(point: list, planes: PlaneSet):
		# The sum of squared distances to the planes
		return sum(plane.distance(point) for plane in planes)

	# Initial point (zero point) for the minimizer
	initial_point = (0, 0, 0)

	min_result = minimize(objective_function, initial_point, args=(planes,), method='L-BFGS-B')
	return min_result.x # closest_point


def closest_point_to_planes_lstsq(planes: PlaneSet):
	A = np.array([[plane.A, plane.B, plane.C] for plane in planes])
	B = np.array([-plane.D for plane in planes])
	X = np.linalg.lstsq(A, B, rcond=None)[0]
	return X


def closest_point_to_planes_pinv(planes: PlaneSet):
	A = np.array([[plane.A, plane.B, plane.C] for plane in planes])
	# A += np.eye(A.shape[0], A.shape[1]) * reg_term
	B = np.array([-plane.D for plane in planes])
	# Compute the pseudo-inverse of A
	A_pinv = np.linalg.pinv(A)
	# Get the least-squares solution to the A system
	return A_pinv.dot(B)


def get_transformed_vertexes_intersect_topology(planes_for_vertex_dict: PlanesForVertex, tr_planes: list[Plane]) -> list[np.array]:
	# Get transformed vertexes by finding closest points to transformed planes
	result_vertexes : list[np.array] = []
	for planes_for_vertex_set in planes_for_vertex_dict.values():
		tr_planes_for_vertex_set: PlaneSet = set()

		# Find all transformed planes (by Plane.id) that will later represent transformed vertex
		for plane_for_vertex in planes_for_vertex_set:
			tr_planes_for_vertex_set.add(tr_planes[plane_for_vertex.id])

		# Find vertex after transformation
		result_vertexes.append(closest_point_to_planes_pinv(tr_planes_for_vertex_set))
	return result_vertexes


def get_transformed_vertexes_sidor_topology(planes_for_vertex_dict: TriPlanesForVertex, tr_planes: list[Plane]) -> list[np.array]:
	# Get transformed vertexes by finding closest points to transformed planes
	result_vertexes : list[np.array] = []
	for tri_plane_for_vertex_set in planes_for_vertex_dict.values():
		tr_sub_vertexes: list[np.array] = list()

		# Find all transformed planes (by Plane.id) that will later represent transformed sub vertex
		for tri_plane_for_vertex in tri_plane_for_vertex_set:
			tr_tri_plane_for_vertex_set = set(
				tr_planes[plane.id] for plane in tri_plane_for_vertex
			)
			tr_sub_vertex = closest_point_to_planes_pinv(tr_tri_plane_for_vertex_set) # planes_intersection
			tr_sub_vertexes.append(tr_sub_vertex)

		# Find mean vertex from list of vertexes
		tr_mean_vertex = np.mean(tr_sub_vertexes, axis=0)
		result_vertexes.append(tr_mean_vertex)

	assert len(result_vertexes) == len(planes_for_vertex_dict)
	return result_vertexes


def get_transformed_vertexes(planes_for_vertex_dict: PlanesForVertex | TriPlanesForVertex, tr_planes: list[Plane], topology: Topology) -> list[np.array]:
	if topology in (topology.Intersect, topology.Orthogonal):
			return get_transformed_vertexes_intersect_topology(planes_for_vertex_dict, tr_planes)
	elif topology == Topology.Sidor:
			return get_transformed_vertexes_sidor_topology(planes_for_vertex_dict, tr_planes)
	assert False


def export_pp_deformed(DEFORMATION_INPUT, DEFORMATION_BASIS_FROM, DEFORMATION_BASIS_TO, DEFORMED_OUTPUT, topology):
	di_vs, di_ts = parse_obj_file(DEFORMATION_INPUT)
	dbf_vs, dbf_ts = parse_obj_file(DEFORMATION_BASIS_FROM)
	dbt_vs, dbt_ts = parse_obj_file(DEFORMATION_BASIS_TO)

	print('low-polygonal models:\n\tdeformation basis from has', len(dbf_ts), 'triangles and \n\tdeformation basis to has', len(dbt_ts), 'triangles.\n')
	assert(len(dbf_ts) == len(dbt_ts))

	elapsed = ElapsedTime()
	in_planes, in_planes_for_vertex_dict = build_planes(di_vs, di_ts, topology)
	exec_stat['plains_build_time'] = elapsed.elapsed()

	# Get transformed planes
	elapsed = ElapsedTime()
	tr_planes = get_polypoint_planes_list(in_planes, orig_basises=dbf_vs, res_basises=dbt_vs)
	tr_vertexes  = get_transformed_vertexes(in_planes_for_vertex_dict, tr_planes, topology)

	exec_stat['total_time'] = elapsed.elapsed()
	#print(f"Transformation took: {total_time} ms")

	# with open(DEFORMED_OUTPUT, 'w+') as f:
	# 	print('exported pp deformed to', DEFORMED_OUTPUT)
	# 	f.write(str_from_vertexes(tr_vertexes) + str_from_faces(di_ts))
	# 	plains_build_time = exec_stat['plains_build_time']
	# 	total_time = exec_stat['total_time']
	# 	f.write(f'\n# plains_build_time: {plains_build_time}')
	# 	f.write(f'\n# total_time: {total_time}\n')


# thorus custom with cube basis
def generate_pp_deformed():
	topo = Topology.Intersect
	topo_str = 'sidor_' if topo == Topology.Sidor else 'ort_' if topo == Topology.Orthogonal else 'intr_'
	DEFORMATION_INPUT = 			'./obj/cube_2/torus_156v.obj'
	DEFORMATION_BASIS_FROM = 		'./obj/cube_2/cube_2.obj'
	DEFORMATION_BASIS_TO = 	        './obj/cube_2/rnd1_1_10/cube_2_rnd1_1.obj'
	DEFORMED_OUTPUT = 				f'./obj/cube_2/rnd1_1_10/thorus_transform/pp_{topo_str}cube_2_rnd1_1.obj'
 
	for v1 in range(1, 11, 1):
		#v1 = round(v1 / 10, 1)
		_DEFORMATION_BASIS_TO = DEFORMATION_BASIS_TO.replace('1.obj', f'{v1}.obj')
		_DEFORMED_OUTPUT = DEFORMED_OUTPUT.replace('1.obj', f'{v1}.obj')
		export_pp_deformed(DEFORMATION_INPUT, DEFORMATION_BASIS_FROM, _DEFORMATION_BASIS_TO, _DEFORMED_OUTPUT, topo)
		print('exported pp deformed with v1', v1, 'DEFORMATION_BASIS_TO:', _DEFORMATION_BASIS_TO)


if __name__ == "__main__":
	# generate_pp_deformed()
 
	topo = Topology.Intersect
	topo_str = 'intr_' if topo == Topology.Sidor else 'ort_' if topo == Topology.Orthogonal else 'intr_'
	DEFORMATION_INPUT = 			'./obj/cube_2/torus_3704v.obj'
	DEFORMATION_BASIS_FROM = 		'./obj/cube_2/cube_2.obj'
	DEFORMATION_BASIS_TO = 	        './obj/cube_2/screwed_1_16/cube_2_screwed_div_5.obj'
	DEFORMED_OUTPUT = 				f'./obj/cube_2/screwed_1_16/torus_3704v_transform/pp_{topo_str}cube_2_screwed_div_5.obj'

	# Run the benchmark
	num_runs = 5
	total_build_planes_time = 0
	total_get_polypoint_planes_list_time = 0
	total_exec_time = 0
	csv_file = './obj/cube_2/screwed_1_16/torus_3704v_transform/benchmark_results.xlsx'

	for _ in range(num_runs):
		export_pp_deformed(DEFORMATION_INPUT, DEFORMATION_BASIS_FROM, DEFORMATION_BASIS_TO, DEFORMED_OUTPUT, topo)
	 
		plains_build_time = exec_stat['plains_build_time']
		get_polypoint_planes_list_time = exec_stat['get_polypoint_planes_list']
		total_time = exec_stat['total_time']
		print(f"Building planes took: {plains_build_time} ms")
		print(f"get_polypoint_planes_list: {get_polypoint_planes_list_time} ms")
		print(f"total_time: {total_time} ms")

		total_build_planes_time += plains_build_time
		total_get_polypoint_planes_list_time += get_polypoint_planes_list_time
		total_exec_time += total_time

	# Calculate averages
	avg_build_planes_time = round(total_build_planes_time / num_runs)
	avg_get_polypoint_planes_list_time = round(total_get_polypoint_planes_list_time / num_runs)
	avg_total_time = round(total_exec_time / num_runs)

	# Check if the XLS file exists and write the header if it doesn't
	try:
		wb = load_workbook(csv_file)
		ws = wb.active
	except FileNotFoundError:
		wb = Workbook()
		ws = wb.active
		ws.append(["Timestamp", "Avg Build Planes Time (ms)", "Avg Get Polypoint Planes List Time (ms)", "Avg Total Time (ms)"])

	# Append results to XLS file
	current_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
	ws.append([current_time, avg_build_planes_time, avg_get_polypoint_planes_list_time, avg_total_time])
	wb.save(csv_file)

	print(f"Averages over {num_runs} runs:")
	print(f"Average building planes time: {avg_build_planes_time} ms")
	print(f"Average get_polypoint_planes_list time: {avg_get_polypoint_planes_list_time} ms")
	print(f"Average total time: {avg_total_time} ms")

